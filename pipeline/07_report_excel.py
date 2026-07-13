"""
07_report_excel.py — ADM + GDZ Ortak STLF Raporu
================================================
Yalnizca kullanici dashboard'da "Musteriye Gonder" dediginde cagrilir.
Forecast ve diagnostic dosyalarini uretmez; iki EDAS icin bes tabloyu yazar.
"""
import sys, os, re, json
import pandas as pd, numpy as np
from pathlib import Path
from datetime import date, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT)); sys.path.insert(0, str(ROOT / "src"))
import config_live as C
from src.output_paths import dated_output_path, glob_output_files, resolve_output_file, DELIVERY_ROOT

# asof_regen.py'nin urettigi REGEN dosyalari her zaman bu sabit (tarih-onde)
# adla yazilir — musteri teslim adlandirmasindan (ADM_forecast_.../GDZ_forecast_...)
# BAGIMSIZDIR, degismedi (bkz. asof_regen.py).
REGEN_FILENAME_TEMPLATE = "{date}_forecast_REGEN.xlsx"

H = list(range(24))
HL = [f"{h:02d}:00" for h in H]
TODAY = date.today()
REPORT = DELIVERY_ROOT / "STLF_LIVE_RAPOR.xlsx"
DEFAULT_REPORT = REPORT
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
HDR = PatternFill("solid", fgColor="4472C4")
HDR_F = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Border(left=Side('thin'),right=Side('thin'),top=Side('thin'),bottom=Side('thin'))

def ga(m, tc, d):
    day = m[m[C.RAW_DATE_COL].dt.date == d]
    if len(day) != 24: return None
    return day.set_index(C.RAW_HOUR_COL)[tc].sort_index()

def lf(p):
    if not p.exists(): return None
    try:
        d = pd.read_excel(p, sheet_name="Tahmin")
        return d.set_index("Saat")["Tahmin_MWh"]
    except: return None

def mape(p, a):
    m = (a>0)&(~np.isnan(p))&(~np.isnan(a))
    return float(np.mean(np.abs((a[m]-p[m])/a[m])*100)) if m.sum() else None

def me(p, a):
    m = (~np.isnan(p))&(~np.isnan(a))
    return float(np.mean(p[m]-a[m])) if m.sum() else None

def make_report(master, tc, fc_dir, sheet_name, filename_template, regen_dir=None):
    """Return 5 DataFrames for the 5 tables.

    fc_dir: canli/musteri teslim forecast'lerinin arandigi kok (DELIVERY_ROOT).
    filename_template: EDAS'a ozgu teslim dosyasi adi (config_live*.OUTPUT_FILENAME_TEMPLATE).
    regen_dir: asof_regen.py'nin urettigi *_REGEN.xlsx dosyalarinin arandigi
               kok — bu dosyalar musteriye gitmiyor, yerel output/'ta kaliyor.
               None ise fc_dir ile ayni (geriye uyum).
    """
    now = TODAY
    ft2 = now + timedelta(days=2); ft1 = now + timedelta(days=1)
    regen_dir = regen_dir if regen_dir is not None else fc_dir

    def forecast_path(d, regen=False):
        if regen:
            return resolve_output_file(regen_dir, REGEN_FILENAME_TEMPLATE.format(date=d))
        return resolve_output_file(fc_dir, filename_template.format(date=d))

    actual_cache = {}
    forecast_cache = {}

    def actual(d):
        if d not in actual_cache:
            actual_cache[d] = ga(master, tc, d)
        return actual_cache[d]

    def forecast(d, regen=False):
        key = (d, regen)
        if key not in forecast_cache:
            forecast_cache[key] = lf(forecast_path(d, regen=regen))
        return forecast_cache[key]

    # Diskte biriken forecast gunleri raporda saga dogru kalici olarak buyur.
    # fc_dir ADM+GDZ ortak (DELIVERY_ROOT); filename_template EDAS'a ozgu literal
    # onek tasidigi icin ("ADM_forecast_"/"GDZ_forecast_") glob asla karismaz.
    date_re = re.compile(r"(\d{4}-\d{2}-\d{2})")
    forecast_dates = []
    for p in glob_output_files(fc_dir, filename_template.format(date="*")):
        m = date_re.search(p.name)
        if m:
            try: forecast_dates.append(date.fromisoformat(m.group(1)))
            except ValueError: pass
    realized = [d for d in forecast_dates if d <= now and actual(d) is not None]
    dates = sorted(set(realized + forecast_dates + [ft1, ft2]))
    fc0 = forecast(TODAY)
    ds = [str(d) for d in dates]
    
    def vcol(cols, rows):
        """rows: list of (hour, [values per date]), cols: [date strings]"""
        out = [["Saat"] + cols + ["ORT"]]
        vals_all = {c: [] for c in cols}
        for h, vals in rows:
            row = [HL[h]]
            nv = []
            for c, v in zip(cols, vals):
                row.append(v if v else "")
                if v: nv.append(v)
            for c, v in zip(cols, vals):
                if v: vals_all[c].append(v)
            row.append(f"{np.nanmean([x for x in vals if x]):.0f}" if any(x for x in vals if x) else "")
            out.append(row)
        # avg row
        avg = ["ORTALAMA"]
        for c in cols:
            v = vals_all.get(c, [])
            avg.append(f"{np.nanmean(v):.0f}" if v else "")
        # total avg
        all_v = [x for lst in vals_all.values() for x in lst]
        avg.append(f"{np.nanmean(all_v):.0f}" if all_v else "")
        out.append(avg)
        return out
    
    # T1: Realized
    t1_rows = []
    for h in H:
        vals = []
        for d in dates:
            if d > now: vals.append(None); continue
            s = actual(d)
            vals.append(s[h] if s is not None and h in s.index and not np.isnan(s[h]) else None)
        t1_rows.append((h, vals))
    t1 = vcol(ds, t1_rows)
    
    # T2: D+1 Forecast
    t2_rows = []
    for h in H:
        vals = []
        for d in dates:
            fc = forecast(d)
            if fc0 is not None and d == ft1: v = fc0.get(h, None)
            elif fc is not None: v = fc.get(h, None)
            else: v = None
            vals.append(v)
        t2_rows.append((h, vals))
    t2 = vcol(ds, t2_rows)
    
    # T3: D+1 Deviation %
    t3_rows = []
    for h in H:
        vals = []
        for d in dates:
            if d > now: vals.append(None); continue
            s = actual(d)
            fc = forecast(d)
            if fc is not None and s is not None and h in fc.index and h in s.index and s[h] > 0:
                vals.append(round(abs(s[h]-fc[h])/s[h]*100, 1))
            else: vals.append(None)
        t3_rows.append((h, vals))
    t3 = vcol(ds, t3_rows)
    # MAPE/ME rows for T3
    mp_row = ["MAPE%"]; me_row = ["ME_MWh"]
    for d in dates:
        if d > now: mp_row.append(""); me_row.append(""); continue
        s = actual(d); fc = forecast(d)
        if fc is not None and s is not None:
            fv = np.array([fc.get(h,np.nan) for h in H]); av = np.array([s.get(h,np.nan) for h in H])
            mp_row.append(f"{mape(fv,av):.2f}" if mape(fv,av) is not None else "")
            me_row.append(f"{me(fv,av):.1f}" if me(fv,av) is not None else "")
        else: mp_row.append(""); me_row.append("")
    mp_row.append(""); me_row.append("")
    t3.append(mp_row); t3.append(me_row)
    
    # T4: D+2 Forecast
    t4_rows = []
    for h in H:
        vals = []
        for d in dates:
            fc = forecast(d, regen=True)
            if fc is None: fc = forecast(d)
            if fc is not None: v = fc.get(h, None)
            elif fc0 is not None and d == ft2: v = fc0.get(h, None)
            else: v = None
            vals.append(v)
        t4_rows.append((h, vals))
    t4 = vcol(ds, t4_rows)
    
    # T5: D+2 Deviation %
    t5_rows = []
    for h in H:
        vals = []
        for d in dates:
            if d > now: vals.append(None); continue
            s = actual(d)
            fc = forecast(d, regen=True)
            if fc is None: fc = forecast(d)
            if fc is not None and s is not None and h in fc.index and h in s.index and s[h] > 0:
                vals.append(round(abs(s[h]-fc[h])/s[h]*100, 1))
            else: vals.append(None)
        t5_rows.append((h, vals))
    t5 = vcol(ds, t5_rows)
    # MAPE/ME
    mp_row2 = ["MAPE%"]; me_row2 = ["ME_MWh"]
    for d in dates:
        if d > now: mp_row2.append(""); me_row2.append(""); continue
        s = actual(d)
        fc = forecast(d, regen=True)
        if fc is None: fc = forecast(d)
        if fc is not None and s is not None:
            fv = np.array([fc.get(h,np.nan) for h in H]); av = np.array([s.get(h,np.nan) for h in H])
            mp_row2.append(f"{mape(fv,av):.2f}" if mape(fv,av) is not None else "")
            me_row2.append(f"{me(fv,av):.1f}" if me(fv,av) is not None else "")
        else: mp_row2.append(""); me_row2.append("")
    mp_row2.append(""); me_row2.append("")
    t5.append(mp_row2); t5.append(me_row2)
    
    return {"T1": t1, "T2": t2, "T3": t3, "T4": t4, "T5": t5}

def write_excel(tables, sheet_name):
    """Bes tabloyu sabit bloklarda, gunleri sutunlarda yazar."""
    import openpyxl
    # Try to load existing
    wb = None
    if REPORT.exists():
        try: wb = openpyxl.load_workbook(REPORT)
        except: pass
    if wb is None: wb = openpyxl.Workbook()
    
    # Get or create sheet
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        default_ws = wb["Sheet"]
        if default_ws.max_row == 1 and default_ws.max_column == 1 and default_ws["A1"].value is None:
            wb.remove(default_ws)
    
    # Baslik + 24 saat + ortalama (+ sapma tablolarinda MAPE/ME) icin
    # bloklar cakismayacak sekilde bosluklu tutulur.
    r0 = {"T1": 1, "T2": 31, "T3": 61, "T4": 95, "T5": 125}
    titles = {
        "T1": "1. Gerçekleşen (MWh)",
        "T2": "2. D+1 Tahmin (MWh)",
        "T3": "3. D+1 Tahmin / Gerçekleşen Sapma (%)",
        "T4": "4. D+2 Tahmin (MWh)",
        "T5": "5. D+2 Tahmin / Gerçekleşen Sapma (%)",
    }
    descriptions = {
        "T1": (
            "GERÇEKLEŞEN\n\nİlgili tarihte ölçülen saatlik dağıtılan enerji "
            "miktarını MWh cinsinden gösterir. Her tarih ayrı bir sütundur."
        ),
        "T2": (
            "D+1 TAHMİN\n\nHedef günden bir gün önce üretilen saatlik talep "
            "tahminlerini gösterir. Değerler MWh cinsindendir."
        ),
        "T3": (
            "D+1 SAPMA\n\nD+1 tahmini ile gerçekleşen değer arasındaki mutlak "
            "yüzdesel sapmayı saat bazında gösterir. Alt satırlarda günlük MAPE "
            "ve ortalama hata (ME) bulunur."
        ),
        "T4": (
            "D+2 TAHMİN\n\nHedef günden iki gün önce üretilen saatlik talep "
            "tahminlerini gösterir. Değerler MWh cinsindendir."
        ),
        "T5": (
            "D+2 SAPMA\n\nD+2 tahmini ile gerçekleşen değer arasındaki mutlak "
            "yüzdesel sapmayı saat bazında gösterir. Alt satırlarda günlük MAPE "
            "ve ortalama hata (ME) bulunur."
        ),
    }

    # Sheet her onayda mevcut tum forecast gunlerinden deterministik yenilenir.
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 3
    table_start_col = 5  # E; A:C aciklama, D gorsel bosluk
    
    for tid, rows in tables.items():
        start_row = r0[tid]
        end_row = start_row + len(rows)

        # Aciklama tek bir Excel hucresidir; A:C ve tum tablo yuksekligi
        # birlestirilerek okunabilir bir bilgi karti gibi gosterilir.
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=3)
        explanation = ws.cell(row=start_row, column=1, value=descriptions[tid])
        explanation.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        explanation.font = Font(bold=True, color="1F1F1F", size=11)
        explanation.fill = PatternFill("solid", fgColor="D9EAF7")
        explanation.border = THIN

        title = ws.cell(row=start_row, column=table_start_col, value=titles[tid])
        title.font = BOLD
        
        for ri, row in enumerate(rows):
            for ci, val in enumerate(row):
                cell = ws.cell(row=start_row + ri + 1, column=table_start_col + ci)
                cell.value = val
                cell.border = THIN
                if ri == 0:
                    cell.fill = HDR; cell.font = HDR_F
                elif val and isinstance(val, str) and val.endswith("%"):
                    try:
                        pct = float(val.rstrip("%"))
                        cell.fill = GREEN if pct < 3 else YELLOW if pct < 6 else RED
                        cell.number_format = '0.0"%"'
                    except: pass
    
    wb.save(REPORT)
    print(f"     Rapor: {REPORT.name}")


def run():
    global REPORT
    print("\n[07] ADM + GDZ ORTAK STLF RAPORU...")
    date_re = re.compile(r"(\d{4}-\d{2}-\d{2})")
    adm_forecasts = glob_output_files(DELIVERY_ROOT, C.OUTPUT_FILENAME_TEMPLATE.format(date="*"))
    dated = []
    for p in adm_forecasts:
        if "_REGEN" in p.name:
            continue
        m = date_re.search(p.name)
        if m:
            dated.append((date.fromisoformat(m.group(1)), p))
    if REPORT == DEFAULT_REPORT and dated:
        target = str(max(d for d, _ in dated))
        REPORT = dated_output_path(DELIVERY_ROOT, target, "STLF_LIVE_RAPOR.xlsx", create=True)
    master = pd.read_parquet(C.MASTER_PARQUET)
    master[C.RAW_DATE_COL] = pd.to_datetime(master[C.RAW_DATE_COL])

    tables = make_report(master, C.RAW_TARGET_COL, DELIVERY_ROOT, "ADM",
                          filename_template=C.OUTPUT_FILENAME_TEMPLATE, regen_dir=C.OUTPUT_DIR)
    write_excel(tables, "ADM")

    # GDZ — ortak rapor iki sheet olmadan basarili sayilmaz.
    try:
        gdz_path = ROOT.parent / "gdz talep" / "live"
        sys.path.insert(0, str(gdz_path))
        import config_live_gdz
        gz = pd.read_parquet(config_live_gdz.GDZ_MASTER_PARQUET)
        gz["Tarih"] = pd.to_datetime(gz["Tarih"])
        gz[C.RAW_HOUR_COL] = gz["Tarih"].dt.hour
        gz = gz.rename(columns={"Tarih": C.RAW_DATE_COL,
                                config_live_gdz.GDZ_RAW_TARGET_COL: "Enerji"})
        gz_tables = make_report(
            gz, "Enerji", DELIVERY_ROOT, "GDZ",
            filename_template=config_live_gdz.OUTPUT_FILENAME_TEMPLATE,
            regen_dir=config_live_gdz.OUTPUT_DIR,
        )
        write_excel(gz_tables, "GDZ")
        print("     [GDZ] eklendi")
    except Exception as e:
        raise RuntimeError(f"GDZ rapor sheet'i olusturulamadi: {e}") from e

    return {"status": "ok", "file": str(REPORT)}

if __name__ == "__main__":
    print(run())
