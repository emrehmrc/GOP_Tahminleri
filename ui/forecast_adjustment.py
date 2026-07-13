"""Diagnostic tabanli forecast inceleme ve kontrollu revizyon arayuzu."""
from __future__ import annotations

import json
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from dashboard_common import LIVE_DIR, GDZ_LIVE_DIR
from output_paths import resolve_output_file, DELIVERY_ROOT
from config_live import OUTPUT_FILENAME_TEMPLATE as ADM_FORECAST_TEMPLATE

if str(GDZ_LIVE_DIR) not in sys.path:
    sys.path.insert(0, str(GDZ_LIVE_DIR))
from config_live_gdz import OUTPUT_FILENAME_TEMPLATE as GDZ_FORECAST_TEMPLATE

DEFAULT_MAX_AI_DELTA_MWH = 250.0
DEFAULT_MAX_AI_HOURS = 24


def _paths(edas: str, target_date: str) -> dict[str, Path]:
    if edas == "ADM":
        return {
            "forecast": resolve_output_file(DELIVERY_ROOT, ADM_FORECAST_TEMPLATE.format(date=target_date)),
            "diagnostic": resolve_output_file(DELIVERY_ROOT, f"diagnostic_{target_date}.html"),
            "diagnostic_json": resolve_output_file(DELIVERY_ROOT, f"diagnostic_{target_date}.json"),
            "diagnostic_script": LIVE_DIR / "pipeline" / "08_diagnostic_html.py",
            "audit_dir": LIVE_DIR / "output" / "adjustments",
        }
    return {
        "forecast": resolve_output_file(DELIVERY_ROOT, GDZ_FORECAST_TEMPLATE.format(date=target_date)),
        "diagnostic": resolve_output_file(DELIVERY_ROOT, f"diagnostic_gdz_{target_date}.html"),
        "diagnostic_json": resolve_output_file(DELIVERY_ROOT, f"diagnostic_gdz_{target_date}.json"),
        "diagnostic_script": GDZ_LIVE_DIR / "pipeline" / "08_diagnostic_html.py",
        "audit_dir": GDZ_LIVE_DIR / "output" / "adjustments",
    }


def _extract_data_from_html(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    marker = "const DATA="
    start = text.find(marker)
    if start < 0:
        return {}
    start += len(marker)
    end = text.find(";\n", start)
    if end < 0:
        end = text.find(";</script>", start)
    return json.loads(text[start:end]) if end >= 0 else {}


def load_diagnostic(edas: str, target_date: str) -> dict:
    p = _paths(edas, target_date)
    if p["diagnostic_json"].exists():
        try:
            return json.loads(p["diagnostic_json"].read_text(encoding="utf-8"))
        except Exception:
            pass
    if p["diagnostic"].exists():
        try:
            return _extract_data_from_html(p["diagnostic"])
        except Exception:
            pass
    return {}


def load_forecast(edas: str, target_date: str) -> pd.DataFrame:
    path = _paths(edas, target_date)["forecast"]
    df = pd.read_excel(path, sheet_name="Tahmin")
    required = {"Tarih", "Saat", "Tahmin_MWh"}
    if not required.issubset(df.columns):
        raise ValueError(f"Forecast semasi gecersiz: eksik={sorted(required - set(df.columns))}")
    df = df.sort_values("Saat").reset_index(drop=True)
    if len(df) != 24 or set(df["Saat"].astype(int)) != set(range(24)):
        raise ValueError("Forecast 0..23 saatlerini tam olarak icermiyor.")
    return df


def sync_diagnostic_with_forecast(data: dict, forecast: pd.DataFrame) -> dict:
    """Diagnostic REC tahminlerini, çıktı Excel'indeki güncel değerlerle değiştir.

    Hava beklentisi/P95/tarihsel örnek gibi diagnostic bulguları korunur; yalnızca
    önerinin başlangıç noktası olan fc değeri güncel müşteri Excel'inden gelir.
    Girdi sözlüğü değiştirilmez.
    """
    synced = dict(data)
    synced_rec = [dict(row) for row in (data.get("REC") or [])]
    values_by_hour = (
        forecast.assign(Saat=forecast["Saat"].astype(int))
        .set_index("Saat")["Tahmin_MWh"]
        .astype(float)
        .to_dict()
    )
    for row in synced_rec:
        try:
            hour = int(row.get("h"))
        except (TypeError, ValueError):
            continue
        if hour in values_by_hour:
            row["fc"] = float(values_by_hour[hour])
    synced["REC"] = synced_rec
    return synced


def build_recommendations(
    data: dict,
    max_delta_mwh: float = DEFAULT_MAX_AI_DELTA_MWH,
    max_hours: int = DEFAULT_MAX_AI_HOURS,
) -> pd.DataFrame:
    """Tum saatleri tarayip aciklanabilir, kullanici-onayli adaylar uret."""
    rows = []
    for r in data.get("REC") or []:
        hour = int(r.get("h", -1))
        fc, exp = r.get("fc"), r.get("exp")
        if fc is None or exp is None:
            continue
        gap = float(exp) - float(fc)
        lo, hi = r.get("lo"), r.get("hi")
        outside = bool((lo is not None and fc < lo) or (hi is not None and fc > hi))
        delta = float(np.clip(gap * 0.45, -float(max_delta_mwh), float(max_delta_mwh)))
        proposed = float(fc) + delta
        n = int(r.get("n") or 0)
        confidence = r.get("confidence")
        raw_confidence_score = r.get("confidence_score")
        if confidence not in {"Yüksek", "Orta", "Düşük"} or raw_confidence_score is None:
            if n < 25:
                confidence, raw_confidence_score = "Düşük", 33.0
            elif outside and n >= 40:
                confidence, raw_confidence_score = "Yüksek", 80.0
            elif outside or abs(gap) >= 70:
                confidence, raw_confidence_score = "Orta", 60.0
            else:
                confidence, raw_confidence_score = "Düşük", 40.0
        confidence_score = float(raw_confidence_score)
        if abs(delta) < 0.01:
            change_text = "değişiklik önerilmiyor"
        else:
            direction = "artış" if delta > 0 else "azalış"
            change_text = f"{abs(delta):.1f} MWh {direction} öneriliyor"
        expectation_direction = "yüksek" if gap > 0 else ("düşük" if gap < 0 else "aynı seviyede")
        drivers = r.get("drivers") or []
        metric_parts = []
        if r.get("r2") is not None: metric_parts.append(f"R²={float(r['r2']):.2f}")
        if r.get("analog_distance") is not None: metric_parts.append(f"benzer-gün uzaklığı={float(r['analog_distance']):.2f}")
        if r.get("model_spread") is not None: metric_parts.append(f"model yayılımı={float(r['model_spread']):.1f} MWh")
        if r.get("mape30") is not None: metric_parts.append(f"saatlik MAPE30=%{float(r['mape30']):.1f}")
        explanation = (
            f"Konsolide beklenti mevcut tahmine göre {abs(gap):.1f} MWh "
            f"{expectation_direction}; {change_text}. "
            + ("Mevcut tahmin P95 bandının dışında. " if outside else "Tahmin P95 bandı içinde. ")
            + f"Güven: {confidence} (%{confidence_score:.0f}); tarihsel örnek: {n}. "
            + (f"Ana etkenler: {', '.join(drivers)}. " if drivers else "")
            + ("; ".join(metric_parts) + "." if metric_parts else "")
        )
        rows.append({
            "Saat": hour, "Eskisi (MWh)": round(float(fc), 2),
            "Yenisi (MWh)": round(proposed, 2), "Değişim (MWh)": round(delta, 2),
            "Beklenti (MWh)": round(float(exp), 2),
            "P95 Alt": round(float(lo), 2) if lo is not None else None,
            "P95 Üst": round(float(hi), 2) if hi is not None else None,
            "Güven": confidence, "Güven Puanı": round(confidence_score, 1),
            "R²": r.get("r2"), "Benzer Gün Uzaklığı": r.get("analog_distance"),
            "Model Yayılımı (MWh)": r.get("model_spread"),
            "MAPE30 (%)": r.get("mape30"), "Bias30 (MWh)": r.get("bias30"),
            "Ana Etkenler": ", ".join(drivers), "Açıklama": explanation,
            "_score": (10000 if outside else 0) + confidence_score * 100 + abs(gap),
        })
    columns = ["Saat", "Eskisi (MWh)", "Yenisi (MWh)", "Değişim (MWh)",
               "Beklenti (MWh)", "P95 Alt", "P95 Üst", "Güven", "Güven Puanı",
               "R²", "Benzer Gün Uzaklığı", "Model Yayılımı (MWh)",
               "MAPE30 (%)", "Bias30 (MWh)", "Ana Etkenler", "Açıklama"]
    if not rows:
        return pd.DataFrame(columns=columns)
    hour_limit = max(1, min(int(max_hours), 24))
    out = pd.DataFrame(rows).nlargest(hour_limit, "_score").drop(columns="_score")
    return out[columns].sort_values("Saat").reset_index(drop=True)


def _write_forecast_atomic(path: Path, values_by_hour: dict[int, float]) -> None:
    wb = load_workbook(path)
    ws = wb["Tahmin"]
    headers = {cell.value: cell.column for cell in ws[1]}
    hour_col, pred_col = headers.get("Saat"), headers.get("Tahmin_MWh")
    if not hour_col or not pred_col:
        wb.close()
        raise ValueError("Tahmin sheet'inde Saat/Tahmin_MWh kolonu bulunamadi.")
    seen = set()
    for row in range(2, ws.max_row + 1):
        hour = int(ws.cell(row=row, column=hour_col).value)
        if hour in values_by_hour:
            value = float(values_by_hour[hour])
            if not np.isfinite(value) or value <= 0:
                wb.close()
                raise ValueError(f"Saat {hour}: gecersiz tahmin {value}")
            ws.cell(row=row, column=pred_col).value = round(value, 2)
            seen.add(hour)
    missing = set(values_by_hour) - seen
    if missing:
        wb.close()
        raise ValueError(f"Excel'de saat bulunamadi: {sorted(missing)}")
    tmp = path.with_name(path.stem + ".tmp.xlsx")
    wb.save(tmp)
    wb.close()
    tmp.replace(path)


def _regenerate_diagnostic(edas: str, target_date: str) -> None:
    p = _paths(edas, target_date)
    result = subprocess.run(
        [sys.executable, str(p["diagnostic_script"])],
        cwd=str(p["diagnostic_script"].parent.parent),
        capture_output=True, text=True, encoding="utf-8", errors="replace", timeout=180,
    )
    if result.returncode != 0:
        raise RuntimeError(result.stdout[-1000:] + result.stderr[-1000:])


def _refresh_stlf_report(edas: str, target_date: str) -> dict:
    """Son forecast Excel'inden ortak raporun ilgili sheet'ini yenile."""
    script = LIVE_DIR / "pipeline" / "07_report_excel.py"
    result = subprocess.run(
        [sys.executable, str(script), "--target", target_date, "--edas", edas],
        cwd=str(LIVE_DIR), capture_output=True, text=True,
        encoding="utf-8", errors="replace", timeout=180,
    )
    if result.returncode != 0:
        raise RuntimeError((result.stdout + result.stderr)[-2000:])
    return {"status": "ok", "edas": edas, "target_date": target_date,
            "stdout": result.stdout[-1000:]}


def apply_adjustments(edas: str, target_date: str, values_by_hour: dict[int, float], source: str) -> dict:
    p = _paths(edas, target_date)
    before_df = load_forecast(edas, target_date)
    before = before_df.set_index(before_df["Saat"].astype(int))["Tahmin_MWh"].astype(float).to_dict()
    changes = []
    for hour, new_value in sorted(values_by_hour.items()):
        old = float(before[int(hour)])
        if abs(float(new_value) - old) > 1e-9:
            changes.append({"hour": int(hour), "before": round(old, 2),
                            "after": round(float(new_value), 2),
                            "delta": round(float(new_value) - old, 2)})
    if not changes:
        return {"status": "no_change", "changes": []}

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    p["audit_dir"].mkdir(parents=True, exist_ok=True)
    backup = p["audit_dir"] / f"{p['forecast'].stem}_{stamp}.xlsx"
    shutil.copy2(p["forecast"], backup)
    _write_forecast_atomic(p["forecast"], {c["hour"]: c["after"] for c in changes})
    audit = {
        "edas": edas, "target_date": target_date, "source": source,
        "applied_at": datetime.now().isoformat(timespec="seconds"),
        "forecast_file": str(p["forecast"]), "backup_file": str(backup), "changes": changes,
    }
    audit_path = p["audit_dir"] / f"{p['forecast'].stem}_{stamp}.json"
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        _regenerate_diagnostic(edas, target_date)
    except Exception as exc:
        audit["diagnostic_warning"] = str(exc)
        audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    try:
        audit["stlf_report"] = _refresh_stlf_report(edas, target_date)
    except Exception as exc:
        audit["stlf_report_warning"] = str(exc)
    audit_path.write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8")
    return {"status": "ok", "changes": changes, "backup": str(backup), "audit": str(audit_path)}


def _rec_frame(data: dict) -> pd.DataFrame:
    return pd.DataFrame([{
        "Saat": r.get("h"), "Tahmin": r.get("fc"), "Konsolide_Beklenti": r.get("exp"),
        "P95_Alt": r.get("lo"), "P95_Ust": r.get("hi"), "Sicaklik": r.get("temp_fc"),
        "GHI": (r.get("weather") or {}).get("ghi"),
        "Bulut": (r.get("weather") or {}).get("cloud"),
        "Nem": (r.get("weather") or {}).get("humidity"),
        "Ruzgar": (r.get("weather") or {}).get("wind"),
        "Yagis": (r.get("weather") or {}).get("precip"),
        "Gecen_Hafta": r.get("lastweek"), "Ornek_n": r.get("n"),
        "Ridge_Beklentisi": r.get("ridge_exp"), "Analog_Beklentisi": r.get("analog_exp"),
        "Model_Medyani": r.get("model_center"), "Bias_Duzeltilmis": r.get("bias_corrected"),
        "R2": r.get("r2"), "Benzer_Gun_Uzakligi": r.get("analog_distance"),
        "Bant_%": r.get("band_pct"), "Model_Yayilimi": r.get("model_spread"),
        "MAPE_7": r.get("mape7"), "Bias_7": r.get("bias7"),
        "MAPE_30": r.get("mape30"), "Bias_30": r.get("bias30"),
        "Guven": r.get("confidence"), "Guven_Puani": r.get("confidence_score"),
        "Ana_Etkenler": ", ".join(r.get("drivers") or []),
    } for r in data.get("REC") or []])


def _render_network_diagnostics(
    target_date: str, data: dict, max_delta_mwh: float, max_hours: int,
) -> pd.DataFrame:
    rec = _rec_frame(data)
    recommendations = build_recommendations(data, max_delta_mwh=max_delta_mwh, max_hours=max_hours)
    tabs = st.tabs(["Özet", "Tahmin Profili", "Hava Benchmark", "P95 / Belirsizlik",
                    "Geçmiş", "Konsolide Faktörler", "Öneriler"])
    with tabs[0]:
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Tahmin tarihi", target_date)
        c2.metric("Diagnostic saat", int(rec["Konsolide_Beklenti"].notna().sum()) if not rec.empty else 0)
        c3.metric("Müdahale adayı", len(recommendations))
        outside = 0 if rec.empty else int(((rec["Tahmin"] < rec["P95_Alt"]) | (rec["Tahmin"] > rec["P95_Ust"])).fillna(False).sum())
        c4.metric("P95 dışı saat", outside)
        st.caption(
            "Beklenti; takvim, geçmiş yük seviyesi, çok-istasyon hava, benzer günler, "
            "model uzlaşısı ve saatlik geçmiş performanstan konsolide edilir."
        )
    with tabs[1]:
        if not rec.empty:
            st.line_chart(rec.set_index("Saat")[["Tahmin", "Konsolide_Beklenti", "Gecen_Hafta"]])
            st.dataframe(rec[["Saat", "Tahmin", "Gecen_Hafta"]], use_container_width=True, hide_index=True)
    with tabs[2]:
        if not rec.empty:
            weather_cols = ["Saat", "Tahmin", "Konsolide_Beklenti", "Sicaklik", "GHI",
                            "Bulut", "Nem", "Ruzgar", "Yagis", "Ornek_n"]
            st.dataframe(rec[weather_cols], use_container_width=True, hide_index=True)
    with tabs[3]:
        if not rec.empty:
            st.line_chart(rec.set_index("Saat")[["Tahmin", "P95_Alt", "P95_Ust"]])
            st.dataframe(rec[["Saat", "Tahmin", "P95_Alt", "P95_Ust"]], use_container_width=True, hide_index=True)
    with tabs[4]:
        cp = data.get("CP") or {}
        hist = pd.DataFrame({k: v.get("load", []) for k, v in cp.items()})
        if not hist.empty:
            hist.index.name = "Saat"
            st.line_chart(hist)
        sim = pd.DataFrame((data.get("DRIFT") or {}).get("sim") or [])
        if not sim.empty:
            st.dataframe(sim.rename(columns={"lbl": "Referans", "mape": "MAPE_%", "corr": "Korelasyon"}), hide_index=True)
    with tabs[5]:
        if not rec.empty:
            factor_cols = [
                "Saat", "Tahmin", "Konsolide_Beklenti", "Ridge_Beklentisi", "Analog_Beklentisi",
                "Model_Medyani", "Bias_Duzeltilmis", "R2", "Benzer_Gun_Uzakligi",
                "Model_Yayilimi", "MAPE_7", "Bias_7", "MAPE_30", "Bias_30",
                "Guven", "Guven_Puani", "Ana_Etkenler",
            ]
            st.dataframe(rec[factor_cols], use_container_width=True, hide_index=True)
            st.caption(
                "Ridge + en yakın 30 analog gün + model medyanı + yeterli geçmiş varsa "
                "bias düzeltmeli tahmin, kalite puanlarıyla ağırlıklandırılır."
            )
    with tabs[6]:
        if recommendations.empty:
            st.success("Yüksek veya orta-yüksek güvenle müdahale gerektiren saat bulunmadı.")
        else:
            st.dataframe(recommendations, use_container_width=True, hide_index=True)
            st.caption(f"Saatlik öneri ±{max_delta_mwh:.0f} MWh, öneri sayısı en fazla {max_hours} saat ile sınırlıdır.")
    return recommendations


def render(target_date: str, customer_sent: bool = False) -> None:
    st.divider()
    st.markdown("### 🩺 Diagnostic'e Göre Düzelt")
    st.caption("Bulguları inceleyin; tahminler yalnızca açık onayınızla değiştirilir.")
    if customer_sent:
        st.warning("Müşteri gönderimi tamamlandığı için tahmin revizyonu kilitlendi.")

    st.session_state.setdefault("diagnostic_excel_refresh", 0)
    if st.button(
        "🔄 Çıktı Excel'lerinden yenile",
        key=f"refresh_diagnostic_excel_{target_date}",
        help="Kaydedilmiş son ADM ve GDZ tahmin Excel'lerini yeniden okur.",
    ):
        st.session_state["diagnostic_excel_refresh"] += 1
        refreshed, warnings = [], []
        for edas in ("ADM", "GDZ"):
            try:
                _refresh_stlf_report(edas, target_date)
                refreshed.append(edas)
            except Exception as exc:
                warnings.append(f"{edas}: {exc}")
        if refreshed:
            st.success(
                "ADM ve GDZ önerileri son çıktı Excel'lerinden yenilendi; "
                f"STLF raporunda güncellenen sheet'ler: {', '.join(refreshed)}."
            )
        if warnings:
            st.warning("STLF raporu tamamen yenilenemedi: " + " | ".join(warnings))
    refresh_version = int(st.session_state["diagnostic_excel_refresh"])
    st.caption(
        "Öneriler konsolide diagnostic ve belirsizlik bulgularını kullanır; Eskisi ve Yenisi "
        "hesapları ise Çıktılar klasöründeki kaydedilmiş son Excel değerlerinden başlar."
    )

    p1, p2 = st.columns(2)
    with p1:
        max_hours = int(st.number_input(
            "Öneri oluşturulacak saat sayısı / EDAS", min_value=1, max_value=24,
            value=DEFAULT_MAX_AI_HOURS, step=1,
        ))
    with p2:
        max_delta_mwh = float(st.number_input(
            "Saatlik azami değişiklik (MWh)", min_value=10, max_value=250,
            value=int(DEFAULT_MAX_AI_DELTA_MWH), step=10,
        ))
    st.caption("Bu parametreler yalnızca öneri listesini sınırlar; hiçbir satır sizin checkbox onayınız olmadan uygulanmaz.")

    all_recs = {}
    network_tabs = st.tabs(["ADM", "GDZ"])
    for tab, edas in zip(network_tabs, ("ADM", "GDZ")):
        with tab:
            data = load_diagnostic(edas, target_date)
            if not data:
                st.warning(f"{edas} diagnostic verisi bulunamadı.")
                all_recs[edas] = pd.DataFrame()
            else:
                try:
                    forecast = load_forecast(edas, target_date)
                    data = sync_diagnostic_with_forecast(data, forecast)
                    source_path = _paths(edas, target_date)["forecast"]
                    modified_at = datetime.fromtimestamp(source_path.stat().st_mtime).strftime("%d.%m.%Y %H:%M:%S")
                    st.caption(f"Güncel tahmin kaynağı: {source_path.name} · Son kayıt: {modified_at}")
                    all_recs[edas] = _render_network_diagnostics(
                        target_date, data, max_delta_mwh=max_delta_mwh, max_hours=max_hours,
                    )
                except Exception as exc:
                    st.error(
                        f"{edas} çıktı Excel'i okunamadı; eski diagnostic değerleriyle "
                        f"öneri gösterilmiyor: {exc}"
                    )
                    all_recs[edas] = pd.DataFrame()

    c1, c2 = st.columns(2)
    with c1:
        if st.button("✨ Diagnostic önerilerini hazırla", disabled=customer_sent, use_container_width=True):
            st.session_state["ai_apply_pending"] = True
    with c2:
        if st.button("✍️ Sonuçları manuel olarak değiştir", disabled=customer_sent, use_container_width=True):
            st.session_state["manual_adjust_open"] = not st.session_state.get("manual_adjust_open", False)

    if st.session_state.get("ai_apply_pending") and not customer_sent:
        st.markdown("#### Diagnostic değişiklik önizlemesi ve satır bazlı onay")
        st.warning(
            "Aşağıdaki tablo yalnızca önizlemedir. Dosyalar henüz değiştirilmedi. "
            "Uygulanmasını istediğiniz her satırın sonundaki Onay kutusunu ayrı ayrı işaretleyin."
        )
        approved = {}
        preview_tabs = st.tabs(["ADM önerileri", "GDZ önerileri"])
        for tab, edas in zip(preview_tabs, ("ADM", "GDZ")):
            with tab:
                recs = all_recs.get(edas, pd.DataFrame())
                if recs.empty:
                    st.info(f"{edas} için seçili parametrelerle öneri bulunmadı.")
                    approved[edas] = recs
                    continue
                preview = recs.copy()
                preview["Onay"] = False
                edited = st.data_editor(
                    preview, hide_index=True, use_container_width=True,
                    disabled=[c for c in preview.columns if c != "Onay"],
                    key=(
                        f"ai_preview_{edas}_{target_date}_{max_hours}_"
                        f"{int(max_delta_mwh)}_{refresh_version}"
                    ),
                    column_config={
                        "Onay": st.column_config.CheckboxColumn(
                            "Onay", help="Yalnızca işaretlenen değişiklik uygulanır.", default=False,
                        ),
                        "Açıklama": st.column_config.TextColumn("Açıklama", width="large"),
                    },
                )
                approved[edas] = edited[edited["Onay"] == True].copy()

        selected_count = sum(len(df) for df in approved.values())
        st.caption(f"Seçilen değişiklik sayısı: {selected_count}")
        if st.button(
            "Seçili değişiklikleri uygula", type="primary",
            disabled=selected_count == 0, key="apply_selected_ai_changes",
        ):
            results = []
            for edas, selected in approved.items():
                if not selected.empty:
                    vals = dict(zip(selected["Saat"].astype(int), selected["Yenisi (MWh)"].astype(float)))
                    results.append((edas, apply_adjustments(edas, target_date, vals, "diagnostic_ai_selected")))
            st.session_state["ai_apply_pending"] = False
            st.success("Seçili öneriler uygulandı: " + ", ".join(f"{e} {len(r['changes'])} saat" for e, r in results))
            st.rerun()

    if st.session_state.get("manual_adjust_open") and not customer_sent:
        st.markdown("#### Manuel saatlik düzenleme")
        manual_tabs = st.tabs(["ADM", "GDZ"])
        for tab, edas in zip(manual_tabs, ("ADM", "GDZ")):
            with tab:
                try:
                    df = load_forecast(edas, target_date)
                except Exception as exc:
                    st.error(str(exc))
                    continue
                edit = df[["Saat", "Tahmin_MWh"]].copy()
                edit["Saat"] = edit["Saat"].map(lambda h: f"{int(h):02d}:00")
                edited = st.data_editor(
                    edit, hide_index=True, use_container_width=True, disabled=["Saat"],
                    key=f"manual_editor_{edas}_{target_date}",
                    column_config={"Tahmin_MWh": st.column_config.NumberColumn(
                        "Tahmin (MWh)", min_value=1.0, step=1.0, format="%.2f")},
                )
                confirm = st.checkbox(f"{edas} manuel değişikliklerini onaylıyorum", key=f"manual_confirm_{edas}")
                if st.button(f"{edas} değişikliklerini kaydet", disabled=not confirm, key=f"manual_save_{edas}"):
                    original = df.set_index(df["Saat"].astype(int))["Tahmin_MWh"].astype(float)
                    new_vals = {h: float(edited.iloc[h]["Tahmin_MWh"]) for h in range(24)}
                    changed = {h: v for h, v in new_vals.items() if abs(v - original.loc[h]) > 1e-9}
                    result = apply_adjustments(edas, target_date, changed, "manual_ui")
                    if result["status"] == "ok":
                        st.success(
                            f"{edas}: {len(result['changes'])} saat güncellendi; "
                            "diagnostic ve STLF raporu yenilendi."
                        )
                        st.rerun()
                    else:
                        st.info("Değişiklik bulunmadı.")
