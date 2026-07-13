from __future__ import annotations

import importlib.util
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook


ROOT = Path(__file__).parents[1]
SPEC = importlib.util.spec_from_file_location(
    "stlf_report_test_module", ROOT / "pipeline" / "07_report_excel.py"
)
REPORT = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(REPORT)


def _write_actual(root: Path, region: str, data_date: date, base: float, comma_date=False) -> None:
    folder_date = data_date + timedelta(days=1)
    folder = root / folder_date.strftime("%Y.%m") / str(folder_date.day)
    folder.mkdir(parents=True, exist_ok=True)
    token = "Aydem" if region == "ADM" else "Gediz"
    separator = "," if comma_date else "."
    rows = ["Asset Id;Starts dd.mm.YYYY HH:MM;Time zone;Energy MWh"]
    for hour in range(24):
        stamp = data_date.strftime(f"%d{separator}%m{separator}%Y") + f" {hour:02d}:00"
        value = f"{base + hour:.2f}".replace(".", ",")
        rows.append(f"DemandaBereket_{token};{stamp};Europe/Istanbul;{value}")
    (folder / f"DemandaBereket_{token}_Daily.csv").write_text("\n".join(rows), encoding="utf-8")


def _write_forecast(root: Path, edas: str, target: date, base: float, old_name=False) -> Path:
    folder = root / target.strftime("%Y.%m") / str(target.day)
    folder.mkdir(parents=True, exist_ok=True)
    name = (
        f"{target}_{edas}_forecast.xlsx" if old_name
        else f"{edas}_forecast_{target}.xlsx"
    )
    path = folder / name
    pd.DataFrame({
        "Tarih": [target] * 24,
        "Saat": list(range(24)),
        "Tahmin_MWh": [base + hour for hour in range(24)],
    }).to_excel(path, sheet_name="Tahmin", index=False)
    return path


def _date_columns(ws, header_row=1):
    return {
        value.date() if hasattr(value, "date") else value: col
        for col in range(3, ws.max_column + 1)
        if (value := ws.cell(header_row, col).value) is not None
    }


def test_three_tables_align_dates_and_compute_requested_deviation(tmp_path):
    inputs, outputs = tmp_path / "talep", tmp_path / "gonderilen"
    d1, d2, d3 = date(2026, 7, 1), date(2026, 7, 2), date(2026, 7, 3)
    for edas in ("ADM", "GDZ"):
        _write_actual(inputs, edas, d1, 100, comma_date=(edas == "GDZ"))
        _write_actual(inputs, edas, d2, 110)
        _write_forecast(outputs, edas, d2, 100, old_name=True)
        _write_forecast(outputs, edas, d3, 120)

    data = {
        edas: REPORT.build_report_data(inputs, outputs, edas, d3)
        for edas in ("ADM", "GDZ")
    }
    path = REPORT.update_workbook(tmp_path / "STLF_LIVE_RAPOR.xlsx", data)
    wb = load_workbook(path, data_only=True)
    assert wb.sheetnames == ["ADM", "GDZ"]
    for ws in wb.worksheets:
        assert {str(rng) for rng in ws.merged_cells.ranges} == {"A8:A10", "A28:A30", "A54:A56"}
        assert ws["B1"].value == ws["B27"].value == ws["B55"].value == "Saat"
        assert ws["B26"].value == ws["B52"].value == ws["B80"].value == "ORTALAMA"
        cols = _date_columns(ws)
        assert cols[d2] == _date_columns(ws, 27)[d2] == _date_columns(ws, 55)[d2]
        col = cols[d2]
        actual = ws.cell(2, col).value
        forecast = ws.cell(28, col).value
        deviation = ws.cell(56, col).value
        assert deviation == pytest.approx(abs(actual / forecast - 1.0))
        assert ws.cell(80, col).number_format == "0.00%"
    wb.close()


def test_rewrite_uses_latest_manually_changed_forecast(tmp_path):
    inputs, outputs = tmp_path / "talep", tmp_path / "gonderilen"
    target = date(2026, 7, 2)
    _write_actual(inputs, "ADM", target, 110)
    forecast = _write_forecast(outputs, "ADM", target, 100)
    report_path = tmp_path / "STLF_LIVE_RAPOR.xlsx"

    first = REPORT.build_report_data(inputs, outputs, "ADM", target)
    REPORT.update_workbook(report_path, {"ADM": first})
    pd.DataFrame({
        "Tarih": [target] * 24,
        "Saat": list(range(24)),
        "Tahmin_MWh": [120 + hour for hour in range(24)],
    }).to_excel(forecast, sheet_name="Tahmin", index=False)
    refreshed = REPORT.build_report_data(inputs, outputs, "ADM", target)
    REPORT.update_workbook(report_path, {"ADM": refreshed})

    wb = load_workbook(report_path, data_only=True)
    ws = wb["ADM"]
    col = _date_columns(ws)[target]
    assert ws.cell(28, col).value == 120
    assert ws.cell(56, col).value == pytest.approx(abs(110 / 120 - 1.0))
    wb.close()
